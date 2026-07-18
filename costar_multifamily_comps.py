"""
costar_multifamily_comps.py

Cleans up to 12 CoStar Unit Mix exports (one per comparable property) into
Unit Type Summary and Comp Summary tabs, mirroring `CoStar/Multifamily
Comps/Reference - Multifamily Market Study Template.xlsx`'s two output
sheets, confirmed by direct inspection (including its two chart objects).

Unit Type ("Br Type"/"Ba Type"/"Unit Type") is derived mechanically from
the raw Beds/Baths columns: Beds == "Studio" -> Br Type="Studio", no Ba
Type, Unit Type="Studio"; numeric Beds -> "{n} Br"/"{n} Ba"/"{Br}/{Ba}".

Rent figures use Effective Rent (falling back to Asking Rent only if a
file doesn't have an Effective column), matching standard real-estate
comp-study practice of comparing what tenants actually pay, not list price.

The reference template's per-comp "%" column in Unit Type Summary turned
out to be a formula bug -- every per-comp row divides by Comp 1's total
unit count regardless of which comp the row is actually on (confirmed by
back-solving the reference numbers: e.g. a 41-unit Palmer Parkside row
shows 0.1289, which is 41/318 -- Courtland Vesta's total, not Palmer
Parkside's own 228). This module computes the intended value instead: a
comp's own row's % is that comp's own share of ITS OWN unit total; the
Total row's % is that unit type's share of the grand total across all
comps.

No engine, no DB -- takes raw upload bytes in per comp, returns an
in-memory openpyxl.Workbook out.
"""

from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BubbleChart, Reference, ScatterChart, Series

from excel_export import write_table

REQUIRED_COLUMNS = ["Beds", "Baths", "Avg SF", "# Units"]

SUMMARY_HEADER = ["#", "%", "Min SF", "Max SF", "Avg SF", "Min Rent", "Max Rent", "Avg Rent", "Min $/SF", "Max $/SF", "Avg $/SF"]


def _unit_type(beds, baths) -> str:
    if str(beds).strip().lower() == "studio":
        return "Studio"
    return f"{beds} Br/{baths} Ba"


def _unit_type_sort_key(t: str):
    if t == "Studio":
        return (0, t)
    try:
        n = int(str(t).split(" ")[0])
    except (ValueError, IndexError):
        n = 99
    return (n, t)


def _find_data_sheet(wb):
    for ws in wb.worksheets:
        if ws.max_row > 1:
            return ws
    raise ValueError("Uploaded file has no data rows")


def parse_unit_mix(file_bytes: bytes) -> list[dict]:
    """Returns raw per-row dicts (unit_type, units, avg_sf, avg_rent,
    avg_rent_sf, units_available), skipping the raw export's own
    pre-aggregated "All Studios"/"Totals" rows -- everything is
    recomputed from the individual unit-type rows instead."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = _find_data_sheet(wb)
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise ValueError(f"Missing expected column(s): {missing}. Is this a CoStar Unit Mix export?")

    rent_col = "Avg Effective Rent/Unit" if "Avg Effective Rent/Unit" in idx else "Avg Asking Rent/Unit"
    rent_sf_col = "Avg Effective Rent/SF" if "Avg Effective Rent/SF" in idx else "Avg Asking Rent/SF"
    if rent_col not in idx or rent_sf_col not in idx:
        raise ValueError("Missing expected rent column(s). Is this a CoStar Unit Mix export?")
    avail_col = "Units Available - Units" if "Units Available - Units" in idx else None

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        beds = raw[idx["Beds"]] if idx["Beds"] < len(raw) else None
        if beds is None:
            continue
        if isinstance(beds, str) and (beds.startswith("All ") or beds.strip() == "Totals"):
            continue
        units = raw[idx["# Units"]] if idx["# Units"] < len(raw) else 0
        if not units:
            continue
        baths = raw[idx["Baths"]] if idx["Baths"] < len(raw) else None
        rows.append({
            "unit_type": _unit_type(beds, baths),
            "units": units,
            "avg_sf": raw[idx["Avg SF"]] or 0,
            "avg_rent": raw[idx[rent_col]] or 0,
            "avg_rent_sf": raw[idx[rent_sf_col]] or 0,
            "units_available": (raw[idx[avail_col]] or 0) if avail_col else 0,
        })
    return rows


def _all_unit_types(parsed: list[dict]) -> list[str]:
    types = {r["unit_type"] for c in parsed for r in c["rows"]}
    return sorted(types, key=_unit_type_sort_key)


def _aggregate(rows: list[dict]) -> dict:
    total_units = sum(r["units"] for r in rows)

    def wavg(key):
        return sum(r[key] * r["units"] for r in rows) / total_units if total_units else 0

    return {
        "count": total_units,
        "min_sf": min((r["avg_sf"] for r in rows), default=0),
        "max_sf": max((r["avg_sf"] for r in rows), default=0),
        "avg_sf": wavg("avg_sf"),
        "min_rent": min((r["avg_rent"] for r in rows), default=0),
        "max_rent": max((r["avg_rent"] for r in rows), default=0),
        "avg_rent": wavg("avg_rent"),
        "min_rent_sf": min((r["avg_rent_sf"] for r in rows), default=0),
        "max_rent_sf": max((r["avg_rent_sf"] for r in rows), default=0),
        "avg_rent_sf": wavg("avg_rent_sf"),
        "units_available": sum(r["units_available"] for r in rows),
    }


def _summary_row(agg: dict, pct: float) -> list:
    return [
        agg["count"], pct, agg["min_sf"], agg["max_sf"], agg["avg_sf"],
        agg["min_rent"], agg["max_rent"], agg["avg_rent"],
        agg["min_rent_sf"], agg["max_rent_sf"], agg["avg_rent_sf"],
    ]


# Unit Type Summary table columns: A=Community+Unit Type, B=Unit Type,
# C=#, D=%, E=Min SF, F=Max SF, G=Avg SF, H=Min Rent, I=Max Rent, ...
_UTS_MIN_SF_COL = 5
_UTS_MAX_SF_COL = 6
_UTS_MIN_RENT_COL = 8
_UTS_MAX_RENT_COL = 9


def _build_unit_type_summary(wb: Workbook, parsed: list[dict]):
    """One block per unit type (Total row + one row per comp that has that
    unit type), a blank row, a range-scatter chart -- same construction as
    Comp Summary's "by Unit Type" chart, but scoped to one unit type at a
    time with each segment representing a comp rather than a unit type
    (confirmed by explicit feedback), then a gap before the next unit
    type's block."""
    ws = wb.create_sheet("Unit Type Summary")
    unit_types = _all_unit_types(parsed)
    grand_total = sum(r["units"] for c in parsed for r in c["rows"])
    header = ["Community + Unit Type", "Unit Type"] + SUMMARY_HEADER

    row_cursor = 1
    for ut in unit_types:
        bucket = [r for c in parsed for r in c["rows"] if r["unit_type"] == ut]
        if not bucket:
            continue
        agg = _aggregate(bucket)
        table = [header, ["Total", ut] + _summary_row(agg, agg["count"] / grand_total if grand_total else 0)]
        comp_first_offset = len(table)  # 0-based offset of the first per-comp row within `table`
        for c in parsed:
            comp_bucket = [r for r in c["rows"] if r["unit_type"] == ut]
            if not comp_bucket:
                continue
            comp_total = sum(r["units"] for r in c["rows"])
            agg2 = _aggregate(comp_bucket)
            table.append([f"{c['name']} {ut}", ut] + _summary_row(agg2, agg2["count"] / comp_total if comp_total else 0))
        end_row = write_table(ws, table, start_row=row_cursor, start_col=1)

        comp_first_row = row_cursor + comp_first_offset
        comp_last_row = end_row - 1
        if comp_last_row >= comp_first_row:
            _add_range_scatter_chart(
                ws, comp_first_row, comp_last_row, anchor_row=end_row + 2, title=f"{ut} -- Comps",
                min_sf_col=_UTS_MIN_SF_COL, max_sf_col=_UTS_MAX_SF_COL,
                min_rent_col=_UTS_MIN_RENT_COL, max_rent_col=_UTS_MAX_RENT_COL,
            )
        row_cursor = end_row + 2 + 20  # blank row + ~20-row chart clearance + blank row before the next unit type


def _build_comp_summary(wb: Workbook, parsed: list[dict]):
    ws = wb.create_sheet("Comp Summary")
    grand_total = sum(r["units"] for c in parsed for r in c["rows"])
    all_rows = [r for c in parsed for r in c["rows"]]
    agg_all = _aggregate(all_rows)

    # --- Block 1: by Community ---
    ws.cell(row=2, column=3, value="Comp Summary by Community")
    header_row = 4
    community_table = [["Comp #", "Community"] + SUMMARY_HEADER + ["Units Available", "% Vacant"]]
    for i, c in enumerate(parsed, start=1):
        agg = _aggregate(c["rows"])
        comp_units = agg["count"]
        pct_vacant = agg["units_available"] / comp_units if comp_units else 0
        community_table.append(
            [f"Comp {i}", c["name"]] + _summary_row(agg, comp_units / grand_total if grand_total else 0) + [agg["units_available"], pct_vacant]
        )
    avg_units_per_comp = grand_total / len(parsed) if parsed else 0
    community_table.append(
        ["All", "Market Average"]
        + [avg_units_per_comp, 1]
        + _summary_row(agg_all, 0)[2:]  # reuse min/max/avg SF+rent+$/SF, pct/count already set above
        + [agg_all["units_available"], agg_all["units_available"] / grand_total if grand_total else 0]
    )
    community_end = write_table(ws, community_table, start_row=header_row, start_col=1)
    community_first_data_row = header_row + 1
    community_last_data_row = community_end - 1  # includes the Market Average row

    # --- Block 2: by Unit Type ---
    unit_types = _all_unit_types(parsed)
    title_row2 = community_end + 2
    ws.cell(row=title_row2, column=4, value="Comp Summary by Unit Type")
    header_row2 = title_row2 + 1
    unit_table = [["Unit Type"] + SUMMARY_HEADER + ["Units Available", "% Vacant"]]
    for ut in unit_types:
        bucket = [r for c in parsed for r in c["rows"] if r["unit_type"] == ut]
        if not bucket:
            continue
        agg = _aggregate(bucket)
        pct_vacant = agg["units_available"] / agg["count"] if agg["count"] else 0
        unit_table.append([ut] + _summary_row(agg, agg["count"] / grand_total if grand_total else 0) + [agg["units_available"], pct_vacant])
    unit_table.append(
        ["All Unit Types"]
        + _summary_row(agg_all, 1)
        + [agg_all["units_available"], agg_all["units_available"] / grand_total if grand_total else 0]
    )
    unit_end = write_table(ws, unit_table, start_row=header_row2, start_col=1)
    unit_first_data_row = header_row2 + 1
    unit_last_data_row = unit_end - 2  # excludes the "All Unit Types" row -- the range chart is per real unit type only

    _add_bubble_chart(ws, community_first_data_row, community_last_data_row, anchor_row=unit_end + 2)
    _add_range_scatter_chart(ws, unit_first_data_row, unit_last_data_row, anchor_row=unit_end + 22, title="Comp Summary by Unit Type")


# Community table columns: A=Comp#, B=Community, C=#, D=%, E=Min SF, F=Max SF,
# G=Avg SF, H=Min Rent, I=Max Rent, J=Avg Rent, K=Min $/SF, L=Max $/SF,
# M=Avg $/SF, N=Units Available, O=% Vacant.
_COMMUNITY_AVG_SF_COL = 7
_COMMUNITY_AVG_RENT_COL = 10
_COMMUNITY_COUNT_COL = 3

# Unit-type table columns: A=Unit Type, B=#, C=%, D=Min SF, E=Max SF, F=Avg SF,
# G=Min Rent, H=Max Rent, I=Avg Rent, ...
_UNIT_MIN_SF_COL = 4
_UNIT_MAX_SF_COL = 5
_UNIT_MIN_RENT_COL = 7
_UNIT_MAX_RENT_COL = 8


def _add_bubble_chart(ws, first_row: int, last_row: int, anchor_row: int):
    """One series per comp (+ the Market Average row) -- matches the
    reference template's construction (each point its own series, for
    per-point coloring), rather than one multi-point series."""
    chart = BubbleChart()
    chart.title = "Comp Summary by Community"
    chart.x_axis.title = "Avg SF"
    chart.y_axis.title = "Avg Rent"
    chart.height, chart.width = 10, 18
    for row in range(first_row, last_row + 1):
        xref = Reference(ws, min_col=_COMMUNITY_AVG_SF_COL, min_row=row, max_row=row)
        yref = Reference(ws, min_col=_COMMUNITY_AVG_RENT_COL, min_row=row, max_row=row)
        zref = Reference(ws, min_col=_COMMUNITY_COUNT_COL, min_row=row, max_row=row)
        chart.series.append(Series(yref, xref, zvalues=zref))
    ws.add_chart(chart, f"A{anchor_row}")


def _add_range_scatter_chart(
    ws, first_row: int, last_row: int, anchor_row: int, title: str,
    min_sf_col: int = _UNIT_MIN_SF_COL, max_sf_col: int = _UNIT_MAX_SF_COL,
    min_rent_col: int = _UNIT_MIN_RENT_COL, max_rent_col: int = _UNIT_MAX_RENT_COL,
):
    """One 2-point line segment per row in [first_row, last_row]: (Min SF,
    Min Rent) -> (Max SF, Max Rent), matching the reference template's
    range-scatter construction exactly. Column positions are parameterized
    (not just the Comp Summary tab's layout) so the same construction can
    be reused on Unit Type Summary -- there, each segment is one comp
    within a given unit type rather than one unit type within the market."""
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "SF"
    chart.y_axis.title = "Rent"
    chart.height, chart.width = 10, 18
    for row in range(first_row, last_row + 1):
        xref = Reference(ws, min_col=min_sf_col, max_col=max_sf_col, min_row=row, max_row=row)
        yref = Reference(ws, min_col=min_rent_col, max_col=max_rent_col, min_row=row, max_row=row)
        chart.series.append(Series(yref, xref))
    ws.add_chart(chart, f"A{anchor_row}")


def build_multifamily_comps_workbook(comps: list[dict]) -> Workbook:
    """comps: [{"name": str, "file_bytes": bytes}], up to 12."""
    parsed = []
    for c in comps:
        rows = parse_unit_mix(c["file_bytes"])
        if not rows:
            raise ValueError(f"No usable unit-type rows found for {c['name']!r}")
        parsed.append({"name": c["name"], "rows": rows})
    if not parsed:
        raise ValueError("No comps provided")

    wb = Workbook()
    wb.remove(wb.active)
    _build_unit_type_summary(wb, parsed)
    _build_comp_summary(wb, parsed)
    return wb
