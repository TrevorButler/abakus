"""
costar_market_overview.py

Cleans CoStar market-metric grid exports for up to 6 named markets and
produces one tab per property class that received at least one upload: a
Year x Market table per metric, plus one line chart per metric (one line
per market that uploaded that class).

Retail/Office/Industrial & Flex genuinely share one raw column format
(confirmed by direct inspection of CommercialDataGrid.xlsx) so one parser
covers all three; Multifamily and Hospitality each have their own raw
format (MultifamilyDataGrid.xlsx / HospitalityDataGrid.xlsx). Only the
confirmed metric subset differs per class:
  - Multifamily (6): Inventory Units, Effective Rent Per Unit, Effective
    Rent Per SF, Vacancy Percent, Absorption Units, Deliveries Units
  - Retail/Office/Industrial & Flex (5, shared): Inventory SF, Vacant
    Percent % Total, Net Absorption SF Total, Deliveries SF, All Service
    Type Rent Overall
  - Hospitality (3): Occupancy, ADR, RevPAR

Hospitality's raw rows are monthly (confirmed: no annual rollup rows exist
in the export, going back decades) while every other class's rows are
already one-per-year -- this module averages by year uniformly for every
class, which is a no-op for the already-annual classes and annualizes
Hospitality the same way the doc frames every other class ("annual market
metrics").

No engine, no DB -- takes raw upload bytes in per market/class, returns an
in-memory openpyxl.Workbook out.
"""

import re
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

from excel_export import unique_sheet_name, write_table

PROPERTY_CLASSES = ["multifamily", "retail", "office", "industrial_flex", "hospitality"]

CLASS_LABELS = {
    "multifamily": "Multifamily",
    "retail": "Retail",
    "office": "Office",
    "industrial_flex": "Industrial & Flex",
    "hospitality": "Hospitality",
}

# (output label, raw column name) pairs -- output label doubles as the
# metric's own sub-table header and chart title.
_COMMERCIAL_METRICS = [
    ("Inventory SF", "Inventory SF"),
    ("Vacant Percent % Total", "Vacant Percent % Total"),
    ("Net Absorption SF Total", "Net Absorption SF Total"),
    ("Deliveries SF", "Deliveries SF"),
    ("All Service Type Rent Overall", "All Service Type Rent Overall"),
]
_MULTIFAMILY_METRICS = [
    ("Inventory Units", "Inventory Units"),
    ("Effective Rent Per Unit", "Effective Rent Per Unit"),
    ("Effective Rent Per SF", "Effective Rent Per SF"),
    ("Vacancy Percent", "Vacancy Percent"),
    ("Absorption Units", "Absorption Units"),
    ("Deliveries Units", "Deliveries Units"),
]
_HOSPITALITY_METRICS = [
    ("Occupancy", "Occupancy"),
    ("ADR", "ADR"),
    ("RevPAR", "RevPAR"),
]

METRICS_BY_CLASS = {
    "retail": _COMMERCIAL_METRICS,
    "office": _COMMERCIAL_METRICS,
    "industrial_flex": _COMMERCIAL_METRICS,
    "multifamily": _MULTIFAMILY_METRICS,
    "hospitality": _HOSPITALITY_METRICS,
}

YEAR_RE = re.compile(r"(\d{4})")


def _find_data_sheet(wb):
    for ws in wb.worksheets:
        if ws.max_row > 1:
            return ws
    raise ValueError("Uploaded file has no data rows")


def _period_year(period):
    m = YEAR_RE.search(str(period)) if period is not None else None
    return int(m.group(1)) if m else None


def parse_grid(file_bytes: bytes, cls: str) -> dict:
    """Returns {metric_label: {year: value}} for one uploaded file."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = _find_data_sheet(wb)
    headers = [str(c.value).strip() if c.value else None for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}
    if "Period" not in idx:
        raise ValueError(f"Missing 'Period' column. Is this a CoStar {CLASS_LABELS[cls]} data grid export?")

    metrics = METRICS_BY_CLASS[cls]
    missing = [raw for _, raw in metrics if raw not in idx]
    if missing:
        raise ValueError(f"Missing expected column(s) {missing} for {CLASS_LABELS[cls]}")

    sums: dict = {label: {} for label, _ in metrics}
    counts: dict = {label: {} for label, _ in metrics}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        year = _period_year(raw[idx["Period"]]) if idx["Period"] < len(raw) else None
        if year is None:
            continue
        for label, col in metrics:
            v = raw[idx[col]] if idx[col] < len(raw) else None
            if v is None:
                continue
            sums[label][year] = sums[label].get(year, 0) + v
            counts[label][year] = counts[label].get(year, 0) + 1

    return {label: {y: sums[label][y] / counts[label][y] for y in sums[label]} for label in sums}


def build_market_overview_workbook(markets: list[dict]) -> Workbook:
    """markets: [{"name": str, "files": {class: bytes}}]. One tab per
    property class that received at least one upload across all markets."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for cls in PROPERTY_CLASSES:
        class_data = {}
        for m in markets:
            file_bytes = m["files"].get(cls)
            if file_bytes is None:
                continue
            class_data[m["name"]] = parse_grid(file_bytes, cls)
        if not class_data:
            continue

        ws = wb.create_sheet(unique_sheet_name(CLASS_LABELS[cls], used_names))
        market_names = list(class_data.keys())
        row_cursor = 1
        for label, _ in METRICS_BY_CLASS[cls]:
            years = sorted({y for series in class_data.values() for y in series.get(label, {})})
            if not years:
                continue
            # None (a true blank cell), not "" -- a market missing a given
            # year is common in real CoStar data (different markets have
            # different history depths), and writing "" instead of a real
            # blank made Excel treat every gap as a hard break rather than
            # skipping over it, chopping each line into disconnected
            # segments (confirmed: this bug didn't show up in testing
            # because the original test fed identical, gap-free data to
            # every market).
            table = [[label], ["Year"] + market_names]
            for y in years:
                table.append([y] + [class_data[m].get(label, {}).get(y) for m in market_names])
            end_row = write_table(ws, table, start_row=row_cursor, start_col=1)

            header_row = row_cursor + 1
            chart = LineChart()
            chart.title = label
            chart.y_axis.title = label
            chart.x_axis.title = "Year"
            chart.height, chart.width = 7, 16
            chart.display_blanks = "span"  # connect the line across a market's missing years instead of breaking
            data_ref = Reference(ws, min_col=2, max_col=1 + len(market_names), min_row=header_row, max_row=header_row + len(years))
            cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(years))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            # Explicit per-series styling so the line itself reads clearly
            # against its markers, and gaps (real blanks now) still draw a
            # continuous line across the missing year rather than stopping.
            for series in chart.series:
                series.smooth = False
                series.marker.symbol = "circle"
                series.marker.size = 5
            ws.add_chart(chart, f"A{end_row + 2}")
            row_cursor = end_row + 2 + 15  # clear the ~15-row-tall chart before the next metric's table

    return wb
