"""
costar_heartbeat.py

Cleans a raw CoStar property-list export into the "Heartbeat" workbook:
per-property broad-class + decade classification, an Existing/Under
Construction/Proposed buildings+RBA summary matrix by class, and a
year-by-class SF-built table feeding one bar chart per class. Mirrors
`CoStar/Heartbeat/CoStar Heartbeat Charts & Development by Decade
Template.xlsx`'s Properties + Heartbeat tabs, confirmed by direct
inspection of that reference workbook.

The reference template's single "New SF Over Time" chart title turned out
to be bound to a selector cell (a dynamic, single chart that redraws for
whichever class is picked elsewhere in the sheet) rather than 7 static
charts -- reproducing that selector mechanism in openpyxl would be fragile
and, for a generated/read-only export, less useful than just shipping one
real chart per class outright. Same reasoning for the year range: the
template hardcodes 1959-2030; this module uses the actual min/max
Year Built found in the upload instead, since a fixed window would be
arbitrary for a general-purpose tool.

No engine, no DB -- takes raw upload bytes in, returns an in-memory
openpyxl.Workbook out (see excel_export.py's module docstring for why
openpyxl specifically: native, editable chart objects).
"""

from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from excel_export import write_table

# 7 broad property classes -- matches the reference template's Properties
# tab column set exactly.
BROAD_CLASSES = ["Retail", "Office", "Multifamily", "Industrial/Flex", "Healthcare", "Hospitality", "Other"]

STATUS_VALUES = ["Existing", "Under Construction", "Proposed"]

# Raw CoStar "Property Type" -> broad class, matched by keyword containment
# rather than exact string equality -- CoStar types often carry a
# parenthetical sub-type (e.g. "Industrial (Industrial Park)", "Industrial
# (Warehouse)"), and exact-match missed those (confirmed against a real
# export). Order matters: checked top to bottom, first match wins.
# Derived from the type variety observed in Reference - CoStar Properties -
# Atlanta Region 2025.xlsx's Raw Properties tab -- flagged for the user's
# review before being treated as final.
PROPERTY_TYPE_KEYWORDS = [
    ("multifamily", "Multifamily"),
    ("multi-family", "Multifamily"),
    ("multi family", "Multifamily"),
    ("apartment", "Multifamily"),
    ("student", "Multifamily"),
    ("health care", "Healthcare"),
    ("healthcare", "Healthcare"),
    ("medical", "Healthcare"),
    ("hospitality", "Hospitality"),
    ("hotel", "Hospitality"),
    ("flex", "Industrial/Flex"),
    ("industrial", "Industrial/Flex"),
    ("warehouse", "Industrial/Flex"),
    ("office", "Office"),
    ("shopping center", "Retail"),
    ("retail", "Retail"),
    ("land", "Other"),
    ("specialty", "Other"),
    ("sports", "Other"),
]

REQUIRED_COLUMNS = ["Property Address", "Property Type", "Building Status", "RBA", "Year Built"]


def broad_class(raw_type) -> str:
    """Anything not matched falls back to 'Other' rather than raising --
    new/unseen CoStar type strings shouldn't hard-fail the whole upload."""
    t = str(raw_type or "").strip().lower()
    for keyword, cls in PROPERTY_TYPE_KEYWORDS:
        if keyword in t:
            return cls
    return "Other"


def _find_data_sheet(wb):
    """CostarExport.xlsx's one data sheet is date-stamped in its own name
    (e.g. 'Export071626') -- take whichever sheet actually has data rows
    rather than matching against a specific sheet name."""
    for ws in wb.worksheets:
        if ws.max_row > 1:
            return ws
    raise ValueError("Uploaded file has no data rows")


def parse_properties(file_bytes: bytes) -> list[dict]:
    """Returns cleaned per-property dicts: address, name, type (broad
    class), status, rba, year_built, decade. Raises ValueError with a
    user-facing message if the upload doesn't look like a CoStar property
    export (missing expected columns)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = _find_data_sheet(wb)
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise ValueError(f"Missing expected column(s): {missing}. Is this a CoStar property export?")

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if idx["Property Address"] >= len(raw) or raw[idx["Property Address"]] is None:
            continue
        year_built = raw[idx["Year Built"]] if idx["Year Built"] < len(raw) else None
        year_built = int(year_built) if isinstance(year_built, (int, float)) else None
        name = raw[idx["Property Name"]] if "Property Name" in idx and idx["Property Name"] < len(raw) else None
        rows.append({
            "address": raw[idx["Property Address"]],
            "name": name or raw[idx["Property Address"]],
            "type": broad_class(raw[idx["Property Type"]]),
            "status": (raw[idx["Building Status"]] or "Existing") if idx["Building Status"] < len(raw) else "Existing",
            "rba": (raw[idx["RBA"]] or 0) if idx["RBA"] < len(raw) else 0,
            "year_built": year_built,
            "decade": (year_built // 10) * 10 if year_built else None,
        })
    return rows


def build_properties_sheet(wb: Workbook, rows: list[dict]):
    """Mirrors the template's Properties tab: one row per property, plus a
    one-hot RBA column per broad class (Total == RBA) for easy SUMIFS-style
    pivoting by whoever opens the download in Excel."""
    ws = wb.create_sheet("Properties")
    header = ["Type", "Name", "Address", "Status", "RBA", "Yr_Built", "Decade"] + BROAD_CLASSES + ["Total"]
    table = [header]
    for r in rows:
        class_cols = [r["rba"] if r["type"] == c else 0 for c in BROAD_CLASSES]
        table.append([r["type"], r["name"], r["address"], r["status"], r["rba"], r["year_built"], r["decade"]] + class_cols + [r["rba"]])
    write_table(ws, table)
    return ws


def _status_matrix(rows: list[dict]) -> dict:
    """{class: {status: {"buildings": n, "sf": n}}}, plus a 'Total' class
    row summed across all classes -- mirrors the template's top-left
    Existing/Under Construction/Proposed summary block exactly."""
    matrix = {c: {s: {"buildings": 0, "sf": 0} for s in STATUS_VALUES} for c in BROAD_CLASSES}
    for r in rows:
        status = r["status"] if r["status"] in STATUS_VALUES else "Existing"
        cell = matrix[r["type"]][status]
        cell["buildings"] += 1
        cell["sf"] += r["rba"]
    return matrix


def _year_class_table(rows: list[dict]) -> tuple[list[int], dict]:
    """(years, {class: {year: sf}}) -- SF newly built (by Year Built) per
    class per year, the source table for the "SF Over Time" charts. Uses
    the actual min/max Year Built present rather than a hardcoded window."""
    years_present = [r["year_built"] for r in rows if r["year_built"]]
    if not years_present:
        return [], {}
    years = list(range(min(years_present), max(years_present) + 1))
    by_class = {c: {y: 0 for y in years} for c in BROAD_CLASSES}
    for r in rows:
        if r["year_built"]:
            by_class[r["type"]][r["year_built"]] += r["rba"]
    return years, by_class


def build_heartbeat_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Heartbeat")

    # Table 1: Existing/Under Construction/Proposed buildings+SF by class.
    matrix = _status_matrix(rows)
    header1 = ["Class"] + [f"{s} {metric}" for s in STATUS_VALUES for metric in ("Buildings", "SF")]
    table1 = [header1]
    totals = {s: {"buildings": 0, "sf": 0} for s in STATUS_VALUES}
    for c in BROAD_CLASSES:
        row = [c]
        for s in STATUS_VALUES:
            cell = matrix[c][s]
            row += [cell["buildings"], cell["sf"]]
            totals[s]["buildings"] += cell["buildings"]
            totals[s]["sf"] += cell["sf"]
        table1.append(row)
    total_row = ["Total"]
    for s in STATUS_VALUES:
        total_row += [totals[s]["buildings"], totals[s]["sf"]]
    table1.append(total_row)
    write_table(ws, table1, start_row=1, start_col=1)

    # Table 2: Year x Class SF-built, starting a few columns over so it
    # doesn't collide with table 1 -- same layout convention as the
    # reference template (two tables side by side on one sheet).
    years, by_class = _year_class_table(rows)
    start_col = len(header1) + 2
    header2 = ["Year"] + BROAD_CLASSES + ["Total"]
    table2 = [header2]
    for y in years:
        row_vals = [by_class[c][y] for c in BROAD_CLASSES]
        table2.append([y] + row_vals + [sum(row_vals)])
    write_table(ws, table2, start_row=1, start_col=start_col)

    # One bar chart per class -- x=Year, y=that class's SF built that year.
    if years:
        year_col = start_col
        chart_anchor_row = len(table2) + 3
        for i, cls in enumerate(BROAD_CLASSES):
            data_col = start_col + 1 + i
            chart = BarChart()
            chart.title = f"{cls} SF Over Time"
            chart.y_axis.title = "SF"
            chart.x_axis.title = "Year"
            chart.height, chart.width = 7, 16
            data_ref = Reference(ws, min_col=data_col, min_row=1, max_row=1 + len(years))
            cats_ref = Reference(ws, min_col=year_col, min_row=2, max_row=1 + len(years))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            anchor_row = chart_anchor_row + (i // 2) * 15
            anchor_col_letter = "A" if i % 2 == 0 else "L"
            ws.add_chart(chart, f"{anchor_col_letter}{anchor_row}")

    return ws


def build_heartbeat_workbook(file_bytes: bytes) -> Workbook:
    rows = parse_properties(file_bytes)
    if not rows:
        raise ValueError("No usable property rows found in the upload")
    wb = Workbook()
    wb.remove(wb.active)
    build_properties_sheet(wb, rows)
    build_heartbeat_sheet(wb, rows)
    return wb
