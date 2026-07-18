"""
dashboard_excel_export.py

Chart-bearing workbook builder for the ACS/BLS "Download Data" retrofit.
Takes a dashboard result dict already shaped by demographics_dashboard.py /
bls_dashboard.py (chart_type: line/stacked_bar/bar/multi_line) and returns
an openpyxl.Workbook with one sheet + one native chart per chart -- same
architecture the CoStar/SmartRE modules already use (see excel_export.py):
no engine, no DB, the caller already ran the query, this only formats the
result. Titles are ported from the frontend's chartMeta.ts/blsChartMeta.ts
(ACS_CHART_TITLES, bls_chart_title()) since those were frontend-only until
now -- the backend needs a title to put on the sheet/chart itself.

Chart-type -> openpyxl mapping, and the block layout (title row + header
row + data rows, chart placed 2 rows below, ~15 rows of clearance before
the next block), reuse exactly what costar_market_overview.py's line/bar
split already learned the hard way: years stay real ints (never strings --
openpyxl always emits <numRef> for category refs regardless of actual cell
type, so a string year under a numeric ref is a real mismatch), missing
values are written as None (a true blank cell, not "" -- "" hard-breaks a
line chart at every gap) with display_blanks="span" so a line still
connects across a gap, and explicit marker styling on every line series.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

from excel_export import unique_sheet_name, write_table

# Mirrors frontend/src/lib/chartMeta.ts's CHART_META (title only -- value
# formatting doesn't apply to a raw Excel number the way it does on screen).
ACS_CHART_TITLES = {
    "population": "Population",
    "households": "Households",
    "housing_units": "Housing Units",
    "housing_unit_occupancy": "Housing Unit Occupancy",
    "housing_unit_type": "Housing Unit Type",
    "housing_unit_type_simplified": "Housing Unit Type (Simplified)",
    "year_built": "Year Built",
    "year_moved_in": "Year Moved In",
    "tenure": "Tenure",
    "median_home_value": "Median Home Value",
    "median_rent": "Median Rent",
    "owner_cost_burden": "Owner Cost Burden",
    "renter_cost_burden": "Renter Cost Burden",
    "age_by_cohort": "Age by Cohort",
    "age_by_cohort_simplified": "Age by Cohort (Simplified)",
    "median_age": "Median Age",
    "race": "Race",
    "hispanic_ethnicity": "Hispanic Ethnicity",
    "household_income": "Household Income",
    "household_income_simplified": "Household Income (Simplified)",
    "median_household_income": "Median Household Income",
    "tenure_by_age_owner": "Tenure by Age -- Owner",
    "tenure_by_age_renter": "Tenure by Age -- Renter",
    "tenure_by_income_owner": "Tenure by Income -- Owner",
    "tenure_by_income_renter": "Tenure by Income -- Renter",
    "household_size": "Household Size",
    "household_type": "Household Type",
}


def acs_chart_title(key: str) -> str:
    return ACS_CHART_TITLES.get(key, key)


def bls_chart_title(key: str, naics_sectors: dict) -> str:
    """Mirrors frontend/src/lib/blsChartMeta.ts's blsChartMeta() -- BLS chart
    keys are dynamic (one employment/wage/avg-pay trend per user-toggled
    sector), so titles are derived, not a static dict."""
    if key == "employment_by_sector":
        return "Employment by Sector"
    if key == "avg_pay_by_sector":
        return "Average Pay by Sector"

    for metric, suffix in (("employment", "Employment"), ("wage", "Total Wages"), ("avg_pay", "Average Annual Pay")):
        prefix = f"{metric}_trend_"
        if key.startswith(prefix):
            code = key[len(prefix):]
            sector = naics_sectors.get(code, code)
            return f"{sector} -- {suffix}"
    return key


def _add_block(ws, row_cursor: int, title: str, series_names: list, rows: list, kind: str) -> int:
    """rows: [[year, v1, v2, ...], ...], columns aligned to series_names.
    kind: 'line' | 'stacked_bar' | 'bar'. Returns the row cursor for the
    next block (past the ~15-row-tall chart)."""
    table = [[title], ["Year"] + series_names] + rows
    end_row = write_table(ws, table, start_row=row_cursor, start_col=1)
    header_row = row_cursor + 1
    n = len(rows)
    data_ref = Reference(ws, min_col=2, max_col=1 + len(series_names), min_row=header_row, max_row=header_row + n)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n)

    if kind == "line":
        chart = LineChart()
        chart.display_blanks = "span"
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for series in chart.series:
            series.smooth = False
            series.marker.symbol = "circle"
            series.marker.size = 5
    else:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked" if kind == "stacked_bar" else "clustered"
        if kind == "stacked_bar":
            chart.overlap = 100
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

    chart.title = title
    chart.x_axis.title = "Year"
    chart.height, chart.width = 7, 16
    ws.add_chart(chart, f"A{end_row + 2}")
    return end_row + 2 + 15


def _write_line(ws, title: str, series: dict):
    years = sorted(series)
    rows = [[y, series.get(y)] for y in years]
    _add_block(ws, 1, title, [title], rows, "line")


def _write_multi_line(ws, title: str, series_by_label: dict):
    labels = list(series_by_label.keys())
    years = sorted({y for s in series_by_label.values() for y in s})
    rows = [[y] + [series_by_label[label].get(y) for label in labels] for y in years]
    _add_block(ws, 1, title, labels, rows, "line")


def _write_categories(ws, title: str, categories: dict, kind: str):
    years = sorted(categories)
    labels = list(dict.fromkeys(label for y in years for label in categories[y]))
    rows = [[y] + [categories[y].get(label) for label in labels] for y in years]
    _add_block(ws, 1, title, labels, rows, kind)


def _write_chart_sheet(wb: Workbook, used_names: set, title: str, chart: dict, view_mode: str = "percent"):
    ws = wb.create_sheet(unique_sheet_name(title, used_names))
    chart_type = chart["chart_type"]
    if chart_type == "line":
        _write_line(ws, title, chart["series"])
    elif chart_type == "multi_line":
        _write_multi_line(ws, title, chart["series_by_label"])
    elif chart_type in ("stacked_bar", "bar"):
        # raw_categories (counts) vs categories (percent shares) -- mirrors
        # the frontend's %/# toggle (ChartViewMode) exactly, since these are
        # the same two fields every category_breakdown()-shaped chart
        # already carries specifically so the export doesn't need a second
        # request. Doesn't apply to line/multi_line charts (those are
        # already a single count/dollar/year value, not a share of
        # something), matching the on-screen toggle's own scope.
        source = chart["raw_categories"] if view_mode == "count" else chart["categories"]
        _write_categories(ws, title, source, chart_type)


def build_dashboard_workbook(dashboard: dict, title_for, view_mode: str = "percent") -> Workbook:
    """dashboard: {chart_key: ChartResult} (single geoid or aggregated
    region -- both share the same shape). title_for(key) -> str. view_mode:
    'percent' or 'count', matching the on-screen %/# toggle (ACS only --
    BLS's charts are all line/multi_line, which this doesn't affect)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()
    for key, chart in dashboard.items():
        if not chart:
            continue
        _write_chart_sheet(wb, used_names, title_for(key), chart, view_mode)
    return wb


def build_multi_geo_dashboard_workbook(dashboard_by_geoid: dict, geo_labels: dict, title_for, view_mode: str = "percent") -> Workbook:
    """dashboard_by_geoid: {geoid: {chart_key: ChartResult}} (one dashboard
    per geography, e.g. Comparative Analysis / Regional Analysis
    "Separated"). geo_labels: {geoid: display label}. title_for(key) -> str.

    'line' gets one combined chart with a series per geography (mirrors
    costar_market_overview.py's one-line-per-market chart). 'multi_line'/
    'stacked_bar'/'bar' get one block per geography instead, stacked on the
    same sheet with spacing (mirrors costar_multifamily_comps.py's
    per-unit-type blocks) -- these already carry their own series dimension
    (sector, or category), so a per-geography chart reads far more clearly
    than trying to force geography in as a second series dimension. This
    matches how the on-screen "Separated" view itself renders: a single
    combined chart for line metrics, one small-multiple per geography for
    everything else.
    """
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()
    geoids = list(dashboard_by_geoid.keys())
    if not geoids:
        return wb

    chart_keys = list(dashboard_by_geoid[geoids[0]].keys())
    for key in chart_keys:
        first = dashboard_by_geoid[geoids[0]].get(key)
        if not first:
            continue
        title = title_for(key)
        chart_type = first["chart_type"]
        ws = wb.create_sheet(unique_sheet_name(title, used_names))

        if chart_type == "line":
            series_by_label = {geo_labels.get(g, g): dashboard_by_geoid[g].get(key, {}).get("series", {}) for g in geoids}
            _write_multi_line(ws, title, series_by_label)
            continue

        row_cursor = 1
        for g in geoids:
            chart = dashboard_by_geoid[g].get(key)
            if not chart:
                continue
            label = geo_labels.get(g, g)
            block_title = f"{title} -- {label}"
            if chart_type == "multi_line":
                labels = list(chart["series_by_label"].keys())
                years = sorted({y for s in chart["series_by_label"].values() for y in s})
                rows = [[y] + [chart["series_by_label"][l].get(y) for l in labels] for y in years]
            else:
                source = chart["raw_categories"] if view_mode == "count" else chart["categories"]
                years = sorted(source)
                labels = list(dict.fromkeys(label for y in years for label in source[y]))
                rows = [[y] + [source[y].get(l) for l in labels] for y in years]
            row_cursor = _add_block(ws, row_cursor, block_title, labels, rows, "line" if chart_type == "multi_line" else chart_type)

    return wb
