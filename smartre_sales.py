"""
smartre_sales.py

Cleans up to 20 SmartRE sales-download files (each capped at ~1,000 rows by
SmartRE's own export limit, confirmed identical schema across split files)
into 4 cuts of sale-price/volume-over-time analysis for a user-chosen
comp set of Subdivisions, plus a combined sales list.

4 cuts (New-construction only, confirmed): Overall + Single-Family/
Townhome/Condo. Resale transactions are excluded entirely before any cut
is built -- comp research only ever looks at a subdivision's New sales,
never Resale (explicit feedback) -- so the New/Resale/Overall dimension
that an earlier 12-cut version had collapses to nothing once every row
reaching a cut is already New. The subdivision picker (list_subdivisions,
below) is NOT filtered this way -- its counts intentionally still span
every Sold transaction, since it's just helping a user judge which
subdivisions have enough volume to be worth picking at all, independent of
what the final comp analysis itself narrows to. Each cut gets a scatter
chart (sale date x price, one point per sale -- no connecting line, this
is a dispersion plot, not a trend line) and a stacked bar chart (year x
price-bin counts).

Only Status == "Sold" rows are used -- "Active" rows are current listings,
not completed sales, and their Price is an asking price rather than a
closing price, so they don't belong in a sale-price/volume analysis.

No dedup across the up-to-20 uploaded files: the same home can legitimately
sell more than once within a multi-year window, so repeat rows for one
address are real, distinct sales, not duplicates (confirmed by the user).

Price bins (fixed, confirmed): under $50K, $50K bands up to $500K, $100K
bands up to $1M, then $1M+.
"""

import csv as csv_module
import re
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series

from excel_export import unique_sheet_name, write_table

REQUIRED_COLUMNS = ["New/ Resale", "Status", "Type", "Subdivision", "Price", "Sqft", "Date"]

TYPE_LABELS = {"SF": "Single-Family", "T": "Townhome", "C": "Condo"}

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}
DATE_RE = re.compile(r"([A-Za-z]+)\s+(\d{4})")

# 4 cuts: (sheet name, Type code filter or None) -- New/Resale is no longer
# a cut dimension since build_sales_analysis_workbook() filters every row to
# New before any cut runs (see module docstring).
CUTS = [
    ("Overall", None),
    ("Single-Family", "SF"),
    ("Townhome", "T"),
    ("Condo", "C"),
]


def _build_price_bins():
    bins = [(0, 50_000, "Under $50K")]
    for lo in range(50_000, 500_000, 50_000):
        bins.append((lo, lo + 50_000, f"${lo // 1000}K-${(lo + 50_000) // 1000}K"))
    for lo in range(500_000, 1_000_000, 100_000):
        bins.append((lo, lo + 100_000, f"${lo // 1000}K-${(lo + 100_000) // 1000}K"))
    bins.append((1_000_000, None, "$1M+"))
    return bins


PRICE_BINS = _build_price_bins()
PRICE_BIN_LABELS = [b[2] for b in PRICE_BINS]


def price_bin_label(price: float) -> str:
    for lo, hi, label in PRICE_BINS:
        if hi is None:
            if price >= lo:
                return label
        elif lo <= price < hi:
            return label
    return PRICE_BINS[0][2]


def _parse_date(value):
    m = DATE_RE.match(str(value).strip()) if value else None
    if not m:
        return None
    month = MONTH_NAMES.get(m.group(1).lower())
    if not month:
        return None
    return date(int(m.group(2)), month, 1)


def _find_data_sheet(wb):
    for ws in wb.worksheets:
        if ws.max_row > 1:
            return ws
    raise ValueError("Uploaded file has no data rows")


def _to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return None


def _extract_rows(file_bytes: bytes) -> list[tuple]:
    """Every row (including any title/header rows) as a plain tuple,
    regardless of whether the upload is .xlsx or a raw CSV -- SmartRE
    ships both (confirmed by the original spec and by a real CSV
    reference file). Detected by content, not filename/extension, since
    the API layer doesn't thread the original filename through. CSV cells
    come back as strings (no native numeric/date typing like openpyxl
    gives xlsx cells) -- callers must coerce."""
    if file_bytes[:2] == b"PK":  # ZIP magic bytes -- xlsx
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = _find_data_sheet(wb)
        return [row for row in ws.iter_rows(values_only=True)]
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return [tuple(row) for row in csv_module.reader(text.splitlines())]


def parse_sales_file(file_bytes: bytes) -> list[dict]:
    """Returns cleaned Sold-only rows. Raises ValueError if the upload
    doesn't look like a SmartRE sales download. Header position varies by
    format -- the xlsx version has SmartRE's own title cell on row 1 and
    the real header on row 2, the CSV version has the header directly on
    row 1 (confirmed against real reference files of both formats) -- so
    this scans the first several rows for whichever one actually has
    every required column, rather than assuming a fixed row.

    A previous version only checked rows 1-2 and only handled .xlsx; a
    real user upload that opened fine but didn't match either assumption
    hit the "missing expected column(s)" error below, which is why this
    scans more rows and both formats now."""
    all_rows = _extract_rows(file_bytes)
    if not all_rows:
        raise ValueError("Uploaded file has no data rows")

    idx = None
    header_row_idx = None
    for i, row in enumerate(all_rows[:5]):
        candidate_idx = {str(v).strip(): j for j, v in enumerate(row) if v not in (None, "")}
        if all(c in candidate_idx for c in REQUIRED_COLUMNS):
            idx = candidate_idx
            header_row_idx = i
            break
    if idx is None:
        raise ValueError("Missing expected column(s). Is this a SmartRE sales download?")

    rows = []
    for raw in all_rows[header_row_idx + 1:]:
        if idx["Status"] >= len(raw) or raw[idx["Status"]] != "Sold":
            continue
        price = _to_number(raw[idx["Price"]]) if idx["Price"] < len(raw) else None
        if price is None:
            continue
        sale_date = _parse_date(raw[idx["Date"]]) if idx["Date"] < len(raw) else None
        rows.append({
            "new_resale": raw[idx["New/ Resale"]] if idx["New/ Resale"] < len(raw) else None,
            "type": raw[idx["Type"]] if idx["Type"] < len(raw) else None,
            "subdivision": str(raw[idx["Subdivision"]]) if idx["Subdivision"] < len(raw) and raw[idx["Subdivision"]] not in (None, "") else "N/a",
            "price": price,
            "sqft": _to_number(raw[idx["Sqft"]]) if idx["Sqft"] < len(raw) else None,
            "sale_date": sale_date,
            "county": raw[idx["County"]] if "County" in idx and idx["County"] < len(raw) else None,
            "zip": raw[idx["Zip"]] if "Zip" in idx and idx["Zip"] < len(raw) else None,
            "high_school": raw[idx["High School"]] if "High School" in idx and idx["High School"] < len(raw) else None,
        })
    return rows


def list_subdivisions(files_bytes: list[bytes]) -> list[dict]:
    """Returns [{"name": str, "count": int}], sorted by count descending.
    A plain alphabetical name list buries real neighborhoods among
    placeholder/junk values SmartRE itself writes when its own address-to-
    subdivision match fails ("0", "N/a", a bare city name, etc.) -- surfacing
    the transaction count up front (and sorting by it) makes it obvious
    which subdivisions are actually worth picking, rather than a user
    unknowingly selecting one with a single transaction."""
    counts: dict = {}
    for fb in files_bytes:
        for r in parse_sales_file(fb):
            counts[r["subdivision"]] = counts.get(r["subdivision"], 0) + 1
    return sorted(({"name": name, "count": c} for name, c in counts.items()), key=lambda x: (-x["count"], x["name"]))


def _filter_rows(rows: list[dict], type_code) -> list[dict]:
    if type_code is None:
        return rows
    return [r for r in rows if r["type"] == type_code]


def _build_cut_sheet(wb: Workbook, name: str, rows: list[dict], used_names: set):
    ws = wb.create_sheet(unique_sheet_name(name, used_names))

    dated_rows = sorted((r for r in rows if r["sale_date"]), key=lambda r: r["sale_date"])
    scatter_table = [["Date", "Price"]] + [[r["sale_date"], r["price"]] for r in dated_rows]
    scatter_end = write_table(ws, scatter_table, start_row=1, start_col=1)

    years = sorted({r["sale_date"].year for r in dated_rows})
    counts = {y: {b: 0 for b in PRICE_BIN_LABELS} for y in years}
    for r in dated_rows:
        counts[r["sale_date"].year][price_bin_label(r["price"])] += 1
    bar_start_col = 4
    bar_table = [["Year"] + PRICE_BIN_LABELS] + [[y] + [counts[y][b] for b in PRICE_BIN_LABELS] for y in years]
    bar_end = write_table(ws, bar_table, start_row=1, start_col=bar_start_col)

    if dated_rows:
        chart1 = ScatterChart()
        chart1.title = f"{name} -- Sale Price Over Time"
        chart1.x_axis.title = "Date"
        chart1.y_axis.title = "Price"
        chart1.height, chart1.width = 10, 18
        xref = Reference(ws, min_col=1, min_row=2, max_row=scatter_end - 1)
        yref = Reference(ws, min_col=2, min_row=2, max_row=scatter_end - 1)
        series = Series(yref, xref, title="Sales")
        series.marker.symbol = "circle"
        series.graphicalProperties.line.noFill = True
        chart1.series.append(series)
        ws.add_chart(chart1, f"A{scatter_end + 2}")

    if years:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.grouping = "stacked"
        chart2.overlap = 100
        chart2.title = f"{name} -- Sales by Year & Price Bin"
        chart2.y_axis.title = "Sales Count"
        chart2.x_axis.title = "Year"
        chart2.height, chart2.width = 10, 18
        data_ref = Reference(ws, min_col=bar_start_col + 1, max_col=bar_start_col + len(PRICE_BIN_LABELS), min_row=1, max_row=bar_end - 1)
        cats_ref = Reference(ws, min_col=bar_start_col, min_row=2, max_row=bar_end - 1)
        chart2.add_data(data_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)
        ws.add_chart(chart2, f"A{scatter_end + 22}")


def build_sales_analysis_workbook(files_bytes: list[bytes], subdivisions: list[str]) -> Workbook:
    all_rows = []
    for fb in files_bytes:
        all_rows.extend(parse_sales_file(fb))
    selected = set(subdivisions)
    rows = [r for r in all_rows if r["subdivision"] in selected]
    # Comp research only ever looks at New-construction sales, never Resale
    # (explicit feedback) -- filtered here, upstream of every cut and the
    # Combined Sales list, not in list_subdivisions()'s picker counts above.
    rows = [r for r in rows if r["new_resale"] == "New"]
    if not rows:
        raise ValueError("No New-construction sold rows found for the selected subdivisions")

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()
    for name, type_code in CUTS:
        _build_cut_sheet(wb, name, _filter_rows(rows, type_code), used_names)

    ws = wb.create_sheet(unique_sheet_name("Combined Sales", used_names))
    header = ["New/Resale", "Type", "Subdivision", "Price", "Sqft", "Date", "County", "Zip", "High School"]
    table = [header] + [
        [r["new_resale"], TYPE_LABELS.get(r["type"], r["type"]), r["subdivision"], r["price"], r["sqft"], r["sale_date"], r["county"], r["zip"], r["high_school"]]
        for r in rows
    ]
    write_table(ws, table)
    return wb
